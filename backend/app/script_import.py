from __future__ import annotations

import csv
import re
from io import StringIO
from pathlib import Path

from .chunking import to_numbered_script

_INDEX_CELL = re.compile(r"^\d{1,4}([\.、.)）])?$")
_HEADER_TOKENS = {
    "#",
    "id",
    "no",
    "n",
    "index",
    "line",
    "text",
    "script",
    "content",
    "english",
    "chinese",
    "note",
    "comment",
    "duration",
    "speaker",
    "role",
    "序号",
    "编号",
    "文案",
    "文本",
    "台词",
    "对白",
    "中文",
    "英文",
    "翻译",
    "备注",
    "时长",
    "角色",
    "说话人",
}


def import_spreadsheet(data: bytes, filename: str = "script.xlsx") -> dict:
    suffix = Path(filename).suffix.lower()
    if suffix in {".csv", ".txt"}:
        rows = _read_csv(data)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _read_xlsx(data)
    elif suffix in {".xls"}:
        raise ValueError("旧版 .xls 请另存为 .xlsx 或 .csv 后再导入")
    else:
        raise ValueError("只支持 Excel（.xlsx）或 CSV")

    cells = _rows_to_cells(rows)
    if not cells:
        raise ValueError("表格里没有可用的文案单元格")
    return {
        "markdown": to_numbered_script(cells),
        "count": len(cells),
        "segments": [{"index": i + 1, "text": text, "chars": len(text)} for i, text in enumerate(cells)],
    }


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return re.sub(r"\s+", " ", text.replace("\n", " "))


def _is_header(value: str) -> bool:
    return value.strip().lower() in _HEADER_TOKENS


def _is_header_row(row: list[str]) -> bool:
    filled = [cell for cell in row if cell]
    if not filled:
        return False
    hits = sum(1 for cell in filled if _is_header(cell))
    return hits >= max(1, (len(filled) + 1) // 2)


def _row_cells(row: list[str]) -> list[str]:
    filled = [(index, cell) for index, cell in enumerate(row) if cell]
    if not filled:
        return []
    if len(filled) >= 2 and filled[0][0] == 0 and _INDEX_CELL.match(filled[0][1]):
        filled = filled[1:]
    return [cell for _, cell in filled if not _is_header(cell)]


def _rows_to_cells(rows: list[list[str]]) -> list[str]:
    start = 1 if rows and _is_header_row(rows[0]) else 0
    cells: list[str] = []
    for row in rows[start:]:
        cells.extend(_row_cells(row))
    return cells


def _read_csv(data: bytes) -> list[list[str]]:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(StringIO(text), dialect)
    return [[_cell_text(cell) for cell in row] for row in reader]


def _read_xlsx(data: bytes) -> list[list[str]]:
    from io import BytesIO

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    rows: list[list[str]] = []
    try:
        for sheet in workbook.worksheets:
            for raw in sheet.iter_rows(values_only=True):
                rows.append([_cell_text(cell) for cell in raw])
    finally:
        workbook.close()
    return rows
